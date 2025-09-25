#!/usr/bin/env python3
from __future__ import annotations
import argparse, csv, re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

GN_REGEX = re.compile(r"\bGN=([A-Za-z0-9\-_.]+)\b")

def norm_subject(s: str) -> str:
    s = s.strip().replace(" ", "_")
    s = re.sub(r"[^\w\-\.]+", "_", s)
    return s

def find_header_info(ws, header_rows: int = 3) -> Tuple[int, Optional[int], List[int], List[str]]:
    """Return (acc_col, desc_col, sample_cols, sample_names) using first 3 rows."""
    rows = list(ws.iter_rows(min_row=1, max_row=header_rows, values_only=True))
    if len(rows) < 3:
        raise RuntimeError("Expected at least 3 header rows.")
    row1, row2, row3 = rows[0], rows[1], rows[2]
    acc_col, desc_col = None, None
    for j, v in enumerate(row3):
        val = (str(v).strip() if v is not None else "").lower()
        if val == "accession": acc_col = j
        if val == "description": desc_col = j
    if acc_col is None:
        # fallback check on row2 (rare)
        for j, v in enumerate(row2):
            val = (str(v).strip() if v is not None else "").lower()
            if val == "accession": acc_col = j
            if val == "description": desc_col = j
    if acc_col is None:
        raise RuntimeError("Could not find 'Accession' column in header row 3.")
    sample_cols, sample_names = [], []
    for j in range(len(row3)):
        meas = str(row3[j]).strip().lower() if row3[j] is not None else ""
        run  = str(row2[j]).strip() if row2[j] is not None else ""
        if ("area" in meas) and run:
            sample_cols.append(j)
            sample_names.append(norm_subject(run))
    if not sample_cols:
        raise RuntimeError("No sample columns detected (no 'Area' labels in row 3).")
    return acc_col, desc_col, sample_cols, sample_names

def build_uniprot_to_symbol(mitocarta_csv: str,
                            uni_col_hint: Optional[str] = None,
                            sym_col_hint: Optional[str] = None) -> Dict[str,str]:
    mapping: Dict[str,str] = {}
    with open(mitocarta_csv, "r", newline="", encoding="utf-8") as fh:
        rdr = csv.reader(fh)
        header = next(rdr)
        def pick_idx(hdr: List[str], hint: Optional[str], keys: List[str]) -> int:
            if hint:
                for i,h in enumerate(hdr):
                    if h == hint or h.lower() == hint.lower(): return i
            for i,h in enumerate(hdr):
                if any(k in h.lower() for k in keys): return i
            return -1
        uni_idx = pick_idx(header, uni_col_hint, ["uniprot","primaryaccession","accession","acc"])
        sym_idx = pick_idx(header, sym_col_hint, ["symbol","gene symbol","genesymbol"])
        if uni_idx < 0 or sym_idx < 0:
            raise RuntimeError("Could not find UniProt and Symbol columns in MitoCarta CSV.")
        for row in rdr:
            if not row or len(row) <= max(uni_idx, sym_idx): continue
            u = (row[uni_idx] or "").strip()
            s = (row[sym_idx] or "").strip().upper()
            if not u or not s: continue
            for uid in re.split(r"[;,\s]+", u):
                uid = uid.strip()
                if uid:
                    mapping[uid] = s
    return mapping

def accession_to_symbol(acc: str, desc: Optional[str], use_gn: bool,
                        u2s: Optional[Dict[str,str]]) -> Optional[str]:
    if use_gn and desc:
        m = GN_REGEX.search(desc)
        if m: return m.group(1).upper()
    if u2s is not None:
        if acc in u2s: return u2s[acc]
        m = re.match(r".*\|([OPQ][0-9][A-Z0-9]{3}[0-9])\|.*", acc)
        if m and m.group(1) in u2s: return u2s[m.group(1)]
        for tok in re.split(r"[;,\s|]+", acc):
            if tok in u2s: return u2s[tok]
    return None

def convert_human_reports_xlsx(xlsx_path: str, out_csv: str, sheet: str = "proteins",
                               use_gn: bool = True,
                               mitocarta_csv: Optional[str] = None,
                               mitocarta_uni_col: Optional[str] = None,
                               mitocarta_symbol_col: Optional[str] = None,
                               min_abundance: float = 0.0) -> Tuple[int,int,int]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    acc_col, desc_col, sample_cols, sample_names = find_header_info(ws, header_rows=3)
    u2s = None
    if not use_gn:
        if mitocarta_csv is None:
            raise RuntimeError("Mapping requested but --mitocarta not provided.")
        u2s = build_uniprot_to_symbol(mitocarta_csv, mitocarta_uni_col, mitocarta_symbol_col)
    outp = Path(out_csv); outp.parent.mkdir(parents=True, exist_ok=True)
    with open(outp, "w", newline="", encoding="utf-8") as fh_out:
        w = csv.writer(fh_out)
        w.writerow(["subject_id","protein","abundance"])
        nrows, proteins, subjects = 0, set(), set(sample_names)
        for row in ws.iter_rows(min_row=4, values_only=True):
            acc = row[acc_col]
            if acc is None: continue
            acc = str(acc).strip()
            desc = (str(row[desc_col]).strip() if (desc_col is not None and row[desc_col] is not None) else None)
            sym = accession_to_symbol(acc, desc, use_gn=use_gn, u2s=u2s)
            if sym is None: continue
            for j, subj in zip(sample_cols, sample_names):
                val = row[j]
                if val is None: continue
                try:
                    x = float(val)
                except Exception:
                    continue
                if x <= min_abundance: continue
                w.writerow([subj, sym, f"{x:.6g}"])
                nrows += 1
                proteins.add(sym)
    return nrows, len(subjects), len(proteins)

def main():
    ap = argparse.ArgumentParser(description="Convert Human512Reports.xlsx -> subject_id,protein,abundance")
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--sheet", default="proteins")
    ap.add_argument("--use-gn", action="store_true", help="Use GN=SYMBOL from Description (no MitoCarta needed)")
    ap.add_argument("--mitocarta", default=None, help="MitoCarta CSV (required if not using --use-gn)")
    ap.add_argument("--mitocarta-uni-col", default=None)
    ap.add_argument("--mitocarta-symbol-col", default=None)
    ap.add_argument("--min-abundance", type=float, default=0.0)
    args = ap.parse_args()
    nrows, nsubj, nprot = convert_human_reports_xlsx(
        xlsx_path=args.xlsx, out_csv=args.out, sheet=args.sheet,
        use_gn=args.use_gn, mitocarta_csv=args.mitocarta,
        mitocarta_uni_col=args.mitocarta_uni_col, mitocarta_symbol_col=args.mitocarta_symbol_col,
        min_abundance=args.min_abundance
    )
    print(f"[OK] wrote {args.out}  rows={nrows}  subjects={nsubj}  proteins={nprot}")

if __name__ == "__main__":
    main()
